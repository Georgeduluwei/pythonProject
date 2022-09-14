import time

import openpyxl

from ddt import ddt, data, unpack
from selenium.webdriver.common.by import By

from demo.page import Test01
from selenium import webdriver



def read_excel():
    workbook = openpyxl.load_workbook("sddt.xlsx")
    sheet = workbook["login"]
    print(sheet.max_row ,sheet.max_column)
    allList = []
    for row in range(2, sheet.max_row +1):
        template = []
        for col in range(1 ,sheet.max_column +1):
            template.append(sheet.cell(row ,col).value)
        allList.append(template)
    print(allList)
    return allList

@ddt
class TestCase01(Test01):
    #    *是拆分列表的
    @data(*read_excel())
    @unpack
    def test01_login(self ,username ,password):
        driver = webdriver.Chrome()
        driver.get("http://centre.ih2ome.cn/")
        driver.find_element(By.ID, "username").send_keys(username)
        driver.find_element(By.ID, "psd").send_keys(password)
        driver.find_element(By.XPATH, "//*[@id='user-func']/label/a").click()
        print("登陆测试")
        # time.sleep(3)


    def test02_uesername(self):
        print("测试用户名")

    def test03_password(self):
        print("测试密码")
